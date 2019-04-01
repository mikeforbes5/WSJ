import requests
import xlrd
import pandas as pd
from openpyxl import load_workbook
import multiprocessing
import urllib3
import socket
from lxml import html
writer = pd.ExcelWriter('Data.xlsx')
book = load_workbook('Data.xlsx')
workbook = xlrd.open_workbook('Data.xlsx')
wbkName = 'Data.xlsx'
wks = book['Sheet1']
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
writer.book = book
worksheet = workbook.sheet_by_index(0)
i = 1
x = 2
z = wks.max_row
while i < z:
    ID = [worksheet.cell_value(i, 0)]
    TICKER = str(ID)
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    sock.bind(("", 57054))
    http = urllib3.PoolManager()
    urllib3.disable_warnings()
    workers = multiprocessing.cpu_count() * 2 + 1
    path = 'C:\\Users\\Michael.Forbes\\PycharmProjects\\Rates\\'
    writer = pd.ExcelWriter('Data.xlsx')
    pd.set_option('display.max_columns', None)
    y = []
    if len(TICKER) == 9:
        urlz = 'https://quotes.wsj.com/etf/' + TICKER[2:7]
        pageContent = requests.get(urlz)
        tree = html.fromstring(pageContent.content)
        try:
            rate = tree.xpath('//*[@id="cr_keystock_drawer"]/div[1]/ul/li[5]/div/span/text()')
            date = tree.xpath('//*[@id="cr_keystock_drawer"]/div[1]/ul/li[5]/div/span/small/text()')
            rates = []
            dates = []
            for r in rate:
                rates.append(r)
            for d in date:
                dates.append(d)
            final = rates[0]
            finals = dates[0]
            print(final)
            print(finals)
            u = final[0]
            q = finals[0]
            wks.cell(row=x, column=2).value = str(final)
            wks.cell(row=x, column=3).value = str(finals)
            book.save('Data.xlsx')
            book.close()
        except:
            pass
    if len(TICKER) == 8:
        urlz = 'https://quotes.wsj.com/etf/' + TICKER[2:6]
        pageContent = requests.get(urlz)
        tree = html.fromstring(pageContent.content)
        try:
            rate = tree.xpath('//*[@id="cr_keystock_drawer"]/div[1]/ul/li[5]/div/span/text()')
            date = tree.xpath('//*[@id="cr_keystock_drawer"]/div[1]/ul/li[5]/div/span/small/text()')
            rates = []
            dates = []
            for r in rate:
                rates.append(r)
            for d in date:
                dates.append(d)
            final = rates[0]
            finals = dates[0]
            print(final)
            print(finals)
            u = final[0]
            q = finals[0]
            wks.cell(row=x, column=2).value = str(final)
            wks.cell(row=x, column=3).value = str(finals)
            book.save('Data.xlsx')
            book.close()
        except:
            pass
    elif len(TICKER) == 7:
        urlz = 'https://quotes.wsj.com/etf/' + TICKER[2:5]
        pageContent = requests.get(urlz)
        tree = html.fromstring(pageContent.content)
        try:
            rate = tree.xpath('//*[@id="cr_keystock_drawer"]/div[1]/ul/li[5]/div/span/text()')
            date = tree.xpath('//*[@id="cr_keystock_drawer"]/div[1]/ul/li[5]/div/span/small/text()')
            rates = []
            dates = []
            for r in rate:
                rates.append(r)
            for d in date:
                dates.append(d)
            final = rates[0]
            finals = dates[0]
            print(final)
            print(finals)
            u = final[0]
            q = finals[0]
            wks.cell(row=x, column=2).value = str(final)
            wks.cell(row=x, column=3).value = str(finals)
            book.save('Data.xlsx')
            book.close()
        except:
            pass
    elif len(TICKER) == 6:
        urlz = 'https://quotes.wsj.com/etf/' + TICKER[2:4]
        pageContent = requests.get(urlz)
        tree = html.fromstring(pageContent.content)
        try:
            rate = tree.xpath('//*[@id="cr_keystock_drawer"]/div[1]/ul/li[5]/div/span/text()')
            date = tree.xpath('//*[@id="cr_keystock_drawer"]/div[1]/ul/li[5]/div/span/small/text()')
            rates = []
            dates = []
            for r in rate:
                rates.append(r)
            for d in date:
                dates.append(d)
            final = rates[0]
            finals = dates[0]
            print(final)
            print(finals)
            u = final[0]
            q = finals[0]
            wks.cell(row=x, column=2).value = str(final)
            wks.cell(row=x, column=3).value = str(finals)
            book.save('Data.xlsx')
            book.close()
        except:
            pass
    i = i + 1
    x = x + 1
